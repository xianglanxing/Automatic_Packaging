'''
Author: error: error: git config user.name & please set dead value or install git && error: git config user.email & please set dead value or install git & please set dead value or install git
Date: 2024-12-31 15:54:24
LastEditors: bobo.bsx 2286362745@qq.com
LastEditTime: 2025-02-08 10:59:43
FilePath: \auto_package\bsx_func\os_func.py
Description: è¿™æ˜¯é»˜è®¤è®¾ç½®,è¯·è®¾ç½®`customMade`, æ‰“å¼€koroFileHeaderæŸ¥çœ‹é…ç½® è¿›è¡Œè®¾ç½®: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
import os
import py7zr
import shutil
import configparser
import openpyxl
import pyzipper
import subprocess
import time
from pathlib import Path

from collections import defaultdict

def compress_folder_to_7z(folder_path, output_file, password):
    """
    å°†æ–‡ä»¶å¤¹å‹ç¼©ä¸º 7z æ ¼å¼å¹¶æ·»åŠ å¯†ç ã€‚

    :param folder_path: è¦å‹ç¼©çš„æ–‡ä»¶å¤¹è·¯å¾„
    :param output_file: è¾“å‡ºçš„ 7z æ–‡ä»¶è·¯å¾„
    :param password: å‹ç¼©å¯†ç 
    """
    # ç¡®ä¿æ–‡ä»¶å¤¹è·¯å¾„æœ‰æ•ˆ
    if not os.path.isdir(folder_path):
        raise ValueError(f"æŒ‡å®šçš„è·¯å¾„ä¸æ˜¯æœ‰æ•ˆæ–‡ä»¶å¤¹: {folder_path}")
    
    # ä½¿ç”¨ py7zr å‹ç¼©æ–‡ä»¶å¤¹
    with py7zr.SevenZipFile(output_file, mode='w', password=password) as archive:
        archive.writeall(folder_path, arcname=os.path.basename(folder_path))
    print(f"æ–‡ä»¶å¤¹ {folder_path} å·²å‹ç¼©ä¸º {output_file}ï¼Œå¹¶æ·»åŠ äº†å¯†ç ã€‚")


def zip_folder_with_password(folder_path, zip_file_path, password):
    """
    å‹ç¼©æ–‡ä»¶å¤¹å¹¶æ·»åŠ å¯†ç ã€‚
    
    :param folder_path: è¦å‹ç¼©çš„æ–‡ä»¶å¤¹è·¯å¾„
    :param zip_file_path: ç”Ÿæˆçš„ zip æ–‡ä»¶è·¯å¾„
    :param password: å‹ç¼©æ–‡ä»¶çš„å¯†ç 
    """
    # å°†å¯†ç è½¬æ¢ä¸ºå­—èŠ‚
    password_bytes = password.encode('utf-8')
    
    with pyzipper.AESZipFile(zip_file_path, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zipf:
        zipf.setpassword(password_bytes)
        zipf.setencryption(pyzipper.WZ_AES)
        
        # éå†æ–‡ä»¶å¤¹å¹¶æ·»åŠ åˆ° zip æ–‡ä»¶
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, folder_path)
                zipf.write(file_path, arcname)

    print(f"å‹ç¼©å®Œæˆï¼Œæ–‡ä»¶ä¿å­˜ä¸º: {zip_file_path}")


def get_field_value(header, data, target_field):
    """
    æ ¹æ®å­—æ®µåç§°ä»æ ‡é¢˜å’Œæ•°æ®ä¸­è·å–å¯¹åº”çš„å€¼ã€‚

    :param header: å…ƒç»„æˆ–åˆ—è¡¨ï¼Œè¡¨ç¤ºæ ‡é¢˜è¡Œ
    :param data: å…ƒç»„æˆ–åˆ—è¡¨ï¼Œè¡¨ç¤ºæ•°æ®è¡Œ
    :param target_field: å­—ç¬¦ä¸²ï¼Œç›®æ ‡å­—æ®µåç§°
    :return: å­—æ®µå¯¹åº”çš„å€¼ï¼ˆå¦‚æœå­—æ®µå­˜åœ¨ï¼‰ï¼Œå¦åˆ™è¿”å› None
    """
    if target_field in header:
        field_index = header.index(target_field)  # æ‰¾åˆ°å­—æ®µå¯¹åº”çš„ç´¢å¼•
        return data[field_index]  # è¿”å›å¯¹åº”çš„å€¼
    else:
        return None


def parse_excel(file_path):
    """
    è§£æ Excel æ•°æ®

    :param file_path: Excel æ–‡ä»¶è·¯å¾„
    :param output_file: è¾“å‡ºæ–‡ä»¶è·¯å¾„
    """
    # æ‰“å¼€å·¥ä½œç°¿
    workbook = openpyxl.load_workbook(file_path)
    data = {}  # ç”¨äºå­˜å‚¨æ‰€æœ‰å·¥ä½œè¡¨çš„æ•°æ®

    # éå†æ‰€æœ‰å·¥ä½œè¡¨
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []

        # éå†æ¯ä¸€è¡Œå¹¶å­˜å‚¨æ•°æ®
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(row)
        
        data[sheet_name] = sheet_data
    
    return data

def modify_onlyone_ini_file(file_path, two_para, dst_str):
    """
    ä¿®æ”¹ ini æ–‡ä»¶å†…å®¹ï¼Œä»…æ ¹æ®é…ç½®é¡¹ï¼ˆKeyï¼‰æŸ¥æ‰¾å¹¶æ›´æ–°å€¼ï¼Œå¿½ç•¥èŠ‚åï¼ˆSectionï¼‰ã€‚
    
    :param file_path: ini æ–‡ä»¶è·¯å¾„
    :param two_para: é…ç½®é¡¹ï¼ˆKeyï¼‰
    :param dst_str: é…ç½®å€¼ï¼ˆValueï¼‰
    """
    config = configparser.RawConfigParser()
    config.optionxform = str  # ä¿ç•™é€‰é¡¹ï¼ˆé”®ï¼‰çš„å¤§å°å†™

    # è¯»å– ini æ–‡ä»¶
    config.read(file_path, encoding="utf-8")

    key_found = False  # è®°å½•æ˜¯å¦æ‰¾åˆ°å¹¶ä¿®æ”¹ Key

    # éå†æ‰€æœ‰ Sectionï¼ŒæŸ¥æ‰¾ `two_para`
    for section in config.sections():
        if two_para in config[section]:
            config[section][two_para] = dst_str
            key_found = True

    # å¦‚æœ `two_para` åœ¨ä»»ä½• Section ä¸­éƒ½ä¸å­˜åœ¨ï¼Œåˆ™æ·»åŠ åˆ° DEFAULT
    if not key_found:
        if "DEFAULT" not in config:
            config["DEFAULT"] = {}
        config["DEFAULT"][two_para] = dst_str

    # ä¿å­˜ä¿®æ”¹åçš„ ini æ–‡ä»¶
    with open(file_path, "w", encoding="utf-8") as configfile:
        config.write(configfile)


def modify_ini_file_old(file_path, one_para, two_para, dst_str):
    """
    ä¿®æ”¹ ini æ–‡ä»¶å†…å®¹ï¼Œæ·»åŠ æˆ–æ›´æ–°é…ç½®ï¼ŒåŒæ—¶ä¿ç•™å¤§å°å†™ã€‚
    Tips: ä¼šå¢åŠ ç©ºæ ¼
    :param file_path: ini æ–‡ä»¶è·¯å¾„
    :param one_para: èŠ‚åï¼ˆSectionï¼‰
    :param two_para: é…ç½®é¡¹ï¼ˆKeyï¼‰
    :param dst_str: é…ç½®å€¼ï¼ˆValueï¼‰
    """
    # ä½¿ç”¨ RawConfigParser ç¦ç”¨å¤§å°å†™è½¬æ¢
    config = configparser.RawConfigParser()
    config.optionxform = str  # ä¿ç•™é€‰é¡¹ï¼ˆé”®ï¼‰çš„å¤§å°å†™

    # è¯»å– ini æ–‡ä»¶
    config.read(file_path, encoding="utf-8")

    # ä¿®æ”¹ç°æœ‰å€¼
    if one_para in config:
        config[one_para][two_para] = dst_str
    else:
        config[one_para] = {two_para: dst_str}

    # ä¿å­˜ä¿®æ”¹åçš„ ini æ–‡ä»¶
    with open(file_path, "w", encoding="utf-8") as configfile:
        config.write(configfile)


def copy_and_remove_folder(source_folder, target_folder):
    """
    å…ˆå°† source_folder å¤åˆ¶åˆ° target_folderï¼Œç„¶ååˆ é™¤ source_folderã€‚

    :param source_folder: æºæ–‡ä»¶å¤¹è·¯å¾„ï¼ˆå¦‚ "C:/PC/K1"ï¼‰
    :param target_folder: ç›®æ ‡æ–‡ä»¶å¤¹è·¯å¾„ï¼ˆå¦‚ "C:/PC/DCDC/K1"ï¼‰
    """
    # ç¡®ä¿æºæ–‡ä»¶å¤¹å­˜åœ¨
    if not os.path.exists(source_folder):
        print(f"âŒ æºæ–‡ä»¶å¤¹ä¸å­˜åœ¨: {source_folder}")
        return

    # ç¡®ä¿ç›®æ ‡çˆ¶ç›®å½•å­˜åœ¨
    os.makedirs(os.path.dirname(target_folder), exist_ok=True)

    # å¤åˆ¶æ–‡ä»¶å¤¹
    try:
        shutil.copytree(source_folder, target_folder)
        print(f"âœ… å¤åˆ¶å®Œæˆ: {source_folder} -> {target_folder}")
    except FileExistsError:
        print(f"âš ï¸ ç›®æ ‡æ–‡ä»¶å¤¹å·²å­˜åœ¨: {target_folder}ï¼Œè·³è¿‡å¤åˆ¶")
    
    # åˆ é™¤æºæ–‡ä»¶å¤¹
    try:
        shutil.rmtree(source_folder)
        print(f"ğŸ—‘ï¸ å·²åˆ é™¤åŸæ–‡ä»¶å¤¹: {source_folder}")
    except PermissionError as e:
        print(f"âŒ åˆ é™¤å¤±è´¥ï¼Œæ–‡ä»¶å¯èƒ½è¢«å ç”¨: {e}")


def modify_ini_file(file_path, one_para, two_para, dst_str):
    """
    ä¿®æ”¹ ini æ–‡ä»¶å†…å®¹ï¼Œæ·»åŠ æˆ–æ›´æ–°é…ç½®ï¼ŒåŒæ—¶ä¿ç•™å¤§å°å†™ã€‚
    Tips: ä¼šåˆ é™¤ç©ºæ ¼
    :param file_path: ini æ–‡ä»¶è·¯å¾„
    :param one_para: èŠ‚åï¼ˆSectionï¼‰
    :param two_para: é…ç½®é¡¹ï¼ˆKeyï¼‰
    :param dst_str: é…ç½®å€¼ï¼ˆValueï¼‰
    """
    # ä½¿ç”¨ RawConfigParser ç¦ç”¨å¤§å°å†™è½¬æ¢
    config = configparser.RawConfigParser()
    config.optionxform = str  # ä¿ç•™é€‰é¡¹ï¼ˆé”®ï¼‰çš„å¤§å°å†™

    # è¯»å– ini æ–‡ä»¶
    config.read(file_path, encoding="utf-8")

    # ä¿®æ”¹ç°æœ‰å€¼
    if one_para in config:
        config[one_para][two_para] = dst_str
    else:
        config[one_para] = {two_para: dst_str}

    # ä¿å­˜ä¿®æ”¹åçš„ ini æ–‡ä»¶
    with open(file_path, "w", encoding="utf-8") as configfile:
        for section in config.sections():
            configfile.write(f"[{section}]\n")
            for key, value in config[section].items():
                configfile.write(f"{key}={value}\n")
            configfile.write("\n")


def parse_ini_file(file_path):
    """
    è§£ææŒ‡å®šçš„ ini æ–‡ä»¶å¹¶è¿”å›å…¶å†…å®¹ã€‚
    
    :param file_path: ini æ–‡ä»¶è·¯å¾„
    :return: ä¸€ä¸ªåŒ…å«æ‰€æœ‰ sections å’Œé”®å€¼å¯¹çš„å­—å…¸
    """

    # ä½¿ç”¨ RawConfigParser ç¦ç”¨å¤§å°å†™è½¬æ¢
    config = configparser.RawConfigParser()
    config.optionxform = str  # ä¿ç•™é€‰é¡¹ï¼ˆé”®ï¼‰çš„å¤§å°å†™
    config.read(file_path, encoding="utf-8")

    config_dict = {}
    for section in config.sections():
        config_dict[section] = dict(config.items(section))
    
    return config_dict


def find_file_in_folder(folder_path, file_name):
    """
    åœ¨æŒ‡å®šæ–‡ä»¶å¤¹ä¸­æŸ¥æ‰¾æ˜¯å¦å­˜åœ¨æŒ‡å®šæ–‡ä»¶åã€‚

    :param folder_path: è¦æœç´¢çš„æ–‡ä»¶å¤¹è·¯å¾„
    :param file_name: è¦æŸ¥æ‰¾çš„æ–‡ä»¶å
    :return: å¦‚æœæ‰¾åˆ°è¿”å›æ–‡ä»¶çš„å®Œæ•´è·¯å¾„ï¼Œå¦åˆ™è¿”å› None
    """
    # éå†æ–‡ä»¶å¤¹ä¸­çš„æ–‡ä»¶
    for root, _, files in os.walk(folder_path):
        if file_name in files:
            return os.path.join(root, file_name)
    return None


def get_subfolder_names(folder_path):
    try:
        # åˆ—å‡ºæ–‡ä»¶å¤¹ä¸‹çš„æ‰€æœ‰å†…å®¹
        all_items = os.listdir(folder_path)
        # è¿‡æ»¤å‡ºå­æ–‡ä»¶å¤¹åç§°
        subfolders = [item for item in all_items if os.path.isdir(os.path.join(folder_path, item))]
        return subfolders
    except Exception as e:
        print(f"Error: {e}")
        return []

def find_strings_in_set(strings, target_set):
    """
    åˆ¤æ–­å­—ç¬¦ä¸²åˆ—è¡¨ä¸­å“ªäº›å­—ç¬¦ä¸²å­˜åœ¨äºé›†åˆä¸­ï¼Œå¹¶è¿”å›å¯¹åº”çš„å­—ç¬¦ä¸²åˆ—è¡¨ã€‚

    :param strings: å­—ç¬¦ä¸²åˆ—è¡¨
    :param target_set: ç›®æ ‡é›†åˆ
    :return: å­˜åœ¨äºé›†åˆä¸­çš„å­—ç¬¦ä¸²åˆ—è¡¨
    """
    return [s for s in strings if s in target_set]


def copy_folder_contents(source_folder, target_folder):
    """
    å°†æºæ–‡ä»¶å¤¹å†…çš„æ‰€æœ‰å†…å®¹å¤åˆ¶åˆ°ç›®æ ‡æ–‡ä»¶å¤¹ä¸­ã€‚
    å¦‚æœç›®æ ‡æ–‡ä»¶å¤¹ä¸­å­˜åœ¨ç›¸åŒçš„æ–‡ä»¶æˆ–å­æ–‡ä»¶å¤¹ï¼Œåˆ™è¦†ç›–ã€‚

    :param source_folder: æºæ–‡ä»¶å¤¹è·¯å¾„
    :param target_folder: ç›®æ ‡æ–‡ä»¶å¤¹è·¯å¾„
    """
    if not os.path.exists(source_folder):
        raise ValueError(f"æºæ–‡ä»¶å¤¹ä¸å­˜åœ¨: {source_folder}")
    
    # ç¡®ä¿ç›®æ ‡æ–‡ä»¶å¤¹å­˜åœ¨
    os.makedirs(target_folder, exist_ok=True)

    # éå†æºæ–‡ä»¶å¤¹çš„å†…å®¹
    for item in os.listdir(source_folder):
        if item == 'Sorting_Log':   # ä¸å¤„ç†è¿™ä¸ªé•¿è·¯å¾„
            continue
        source_item = os.path.join(source_folder, item)
        target_item = os.path.join(target_folder, item)

        if os.path.isdir(source_item):  # å¦‚æœæ˜¯å­æ–‡ä»¶å¤¹
            if os.path.exists(target_item):
                shutil.rmtree(target_item)  # åˆ é™¤å·²å­˜åœ¨çš„ç›®æ ‡æ–‡ä»¶å¤¹
            shutil.copytree(source_item, target_item)  # å¤åˆ¶æ•´ä¸ªæ–‡ä»¶å¤¹
        else:  # å¦‚æœæ˜¯æ–‡ä»¶
            shutil.copy2(source_item, target_item)  # å¤åˆ¶æ–‡ä»¶ï¼ˆè¦†ç›–ï¼‰

        # print(f"å·²å¤åˆ¶: {source_item} -> {target_item}")

def search_files_by_name(directory, search_string):
    # ä¿å­˜åŒ¹é…æ–‡ä»¶çš„è·¯å¾„
    matching_files = []

    # éå†æŒ‡å®šç›®å½•
    for root, _, files in os.walk(directory):
        for file in files:
            # æ£€æŸ¥æ–‡ä»¶åæ˜¯å¦åŒ…å«ç›®æ ‡å­—ç¬¦ä¸²
            if search_string in file:
                matching_files.append(os.path.join(root, file))

    return matching_files

def copy_file_to_folder(source_file, target_folder):
    # ç¡®ä¿ç›®æ ‡æ–‡ä»¶å¤¹å­˜åœ¨ï¼Œä¸å­˜åœ¨åˆ™åˆ›å»º
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)

    # è·å–æºæ–‡ä»¶åå¹¶æ„é€ ç›®æ ‡è·¯å¾„
    file_name = os.path.basename(source_file)
    target_path = os.path.join(target_folder, file_name)

    # æ‹·è´æ–‡ä»¶
    shutil.copy2(source_file, target_path)
    print(f"æ–‡ä»¶å·²æˆåŠŸæ‹·è´åˆ°: {target_path}")

def copy_files(source_folder, target_folder):
    """
    å°†æŒ‡å®šæ–‡ä»¶å¤¹å†…çš„æ‰€æœ‰æ–‡ä»¶å¤åˆ¶åˆ°å¦ä¸€ä¸ªæ–‡ä»¶å¤¹ä¸­ï¼Œè‹¥å·²å­˜åœ¨åˆ™å¼ºåˆ¶è¦†ç›–ã€‚

    :param source_folder: æºæ–‡ä»¶å¤¹è·¯å¾„
    :param target_folder: ç›®æ ‡æ–‡ä»¶å¤¹è·¯å¾„
    """
    if not os.path.exists(source_folder):
        raise ValueError(f"æºæ–‡ä»¶å¤¹ä¸å­˜åœ¨: {source_folder}")
    
    # ç¡®ä¿ç›®æ ‡æ–‡ä»¶å¤¹å­˜åœ¨
    os.makedirs(target_folder, exist_ok=True)

    for root, _, files in os.walk(source_folder):
        for file in files:
            source_file = os.path.join(root, file)
            target_file = os.path.join(target_folder, file)
            
            # å¤åˆ¶æ–‡ä»¶ï¼Œå¼ºåˆ¶è¦†ç›–
            shutil.copy2(source_file, target_file)
            print(f"å·²å¤åˆ¶æ–‡ä»¶: {source_file} -> {target_file}")



def copy_folder(source_folder, destination_folder):
    """
    æ‹·è´æ•´ä¸ªæ–‡ä»¶å¤¹åŠå…¶æ‰€æœ‰å†…å®¹ã€‚
    
    :param source_folder: æºæ–‡ä»¶å¤¹è·¯å¾„
    :param destination_folder: ç›®æ ‡æ–‡ä»¶å¤¹è·¯å¾„
    """
    if not os.path.exists(source_folder):
        print(f"æºæ–‡ä»¶å¤¹ä¸å­˜åœ¨: {source_folder}")
        return

    try:
        shutil.copytree(source_folder, destination_folder)
        print(f"æ–‡ä»¶å¤¹æ‹·è´å®Œæˆï¼Œä» {source_folder} åˆ° {destination_folder}")
    except FileExistsError:
        print(f"ç›®æ ‡æ–‡ä»¶å¤¹å·²å­˜åœ¨: {destination_folder}")
    except Exception as e:
        print(f"æ‹·è´æ–‡ä»¶å¤¹å¤±è´¥: {e}")

def extract_7z(file_path, output_folder):
    """
    è§£å‹ .7z æ–‡ä»¶åˆ°æŒ‡å®šç›®å½•

    :param file_path: è¦è§£å‹çš„ .7z æ–‡ä»¶è·¯å¾„
    :param output_folder: è§£å‹åˆ°çš„ç›®æ ‡æ–‡ä»¶å¤¹
    """
    if not os.path.exists(file_path):
        print(f"æ–‡ä»¶ {file_path} ä¸å­˜åœ¨ã€‚")
        return

    # åˆ›å»ºè¾“å‡ºæ–‡ä»¶å¤¹
    os.makedirs(output_folder, exist_ok=True)

    try:
        with py7zr.SevenZipFile(file_path, mode='r') as archive:
            archive.extractall(path=output_folder)
            print(f"æ–‡ä»¶å·²æˆåŠŸè§£å‹åˆ° {output_folder}")
    except Exception as e:
        print(f"è§£å‹æ–‡ä»¶æ—¶å‡ºé”™: {e}")

def force_copy(src, dst):
    """
    å¼ºåˆ¶å¤åˆ¶æ–‡ä»¶ï¼Œæ— è®ºç›®æ ‡æ–‡ä»¶æ˜¯å¦å­˜åœ¨ã€‚
    :param src: æºæ–‡ä»¶è·¯å¾„
    :param dst: ç›®æ ‡æ–‡ä»¶è·¯å¾„
    """
    try:
        # åˆ›å»ºç›®æ ‡ç›®å½•ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        
        # å¼ºåˆ¶å¤åˆ¶æ–‡ä»¶
        shutil.copy2(src, dst)  # ä½¿ç”¨ copy2 ä¿ç•™æ–‡ä»¶å…ƒæ•°æ®
        print(f"æ–‡ä»¶å·²æˆåŠŸå¤åˆ¶åˆ° {dst}")
    except FileNotFoundError:
        print(f"æºæ–‡ä»¶ {src} ä¸å­˜åœ¨ï¼Œè¯·æ£€æŸ¥è·¯å¾„ï¼")
    except Exception as e:
        print(f"å¤åˆ¶æ–‡ä»¶æ—¶å‘ç”Ÿé”™è¯¯ï¼š{e}")


def create_directory(path):
    """
    åˆ›å»ºæ–‡ä»¶å¤¹ï¼Œå¦‚æœæ–‡ä»¶å¤¹å·²å­˜åœ¨åˆ™ä¸é‡å¤åˆ›å»ºã€‚

    :param path: æ–‡ä»¶å¤¹è·¯å¾„ï¼ˆç»å¯¹è·¯å¾„æˆ–ç›¸å¯¹è·¯å¾„ï¼‰
    :return: ä¸€ä¸ªå­—å…¸ï¼ŒåŒ…å«æ–‡ä»¶å¤¹è·¯å¾„åŠçŠ¶æ€
             {"status": "created" or "exists", "path": "folder_path"}
    """
    try:
        if not os.path.exists(path):
            os.makedirs(path)
            return {"status": "created", "path": path}
        else:
            return {"status": "exists", "path": path}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def remove_empty_folders(path):
    """
    æ£€æŸ¥è·¯å¾„ä¸‹çš„æ‰€æœ‰æ–‡ä»¶å¤¹ï¼Œå¦‚æœæ–‡ä»¶å¤¹ä¸ºç©ºåˆ™åˆ é™¤ã€‚
    
    :param path: è¦æ£€æŸ¥çš„æ ¹ç›®å½•è·¯å¾„
    """
    for root, dirs, files in os.walk(path, topdown=False):  # ä»å­ç›®å½•å‘ä¸Šéå†
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            if not os.listdir(dir_path):  # æ£€æŸ¥æ–‡ä»¶å¤¹æ˜¯å¦ä¸ºç©º
                try:
                    os.rmdir(dir_path)  # åˆ é™¤ç©ºæ–‡ä»¶å¤¹
                    print(f"å·²åˆ é™¤ç©ºæ–‡ä»¶å¤¹: {dir_path}")
                except Exception as e:
                    print(f"åˆ é™¤æ–‡ä»¶å¤¹ {dir_path} æ—¶å‡ºé”™: {e}")

def parse_and_group_by_filename(file_path):
    """
    è§£æ Excel è¡¨æ ¼æ•°æ®ï¼Œå¹¶æŒ‰æ–‡ä»¶ååˆ†ç»„ã€‚
    
    :param file_path: Excel æ–‡ä»¶è·¯å¾„
    :return: æŒ‰æ–‡ä»¶ååˆ†ç»„çš„å­—å…¸
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # è·å–æ ‡é¢˜è¡Œ
    headers = [cell.value for cell in sheet[1]]  # ç¬¬ä¸€è¡Œæ ‡é¢˜

    # åˆå§‹åŒ–åˆ†ç»„ç»“æœ
    grouped_data = defaultdict(list)

    # éå†æ•°æ®è¡Œ
    for row in sheet.iter_rows(min_row=2, values_only=True):  # ä»ç¬¬äºŒè¡Œå¼€å§‹
        row_data = dict(zip(headers, row))  # å°†è¡Œæ•°æ®ä¸è¡¨å¤´æ˜ å°„æˆå­—å…¸

        # æ ¹æ®æ–‡ä»¶ååˆ†ç»„
        filename = row_data.get("æ–‡ä»¶å")  # è·å– "æ–‡ä»¶å" åˆ—çš„å€¼
        if filename:  # ç¡®ä¿æ–‡ä»¶åä¸ä¸ºç©º
            grouped_data[filename].append(row_data)

    return grouped_data


def delete_file(file_path):
    """
    åˆ é™¤æŒ‡å®šè·¯å¾„çš„æ–‡ä»¶ã€‚

    :param file_path: æ–‡ä»¶è·¯å¾„ï¼ˆç»å¯¹è·¯å¾„æˆ–ç›¸å¯¹è·¯å¾„ï¼‰
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)  # åˆ é™¤æ–‡ä»¶
            print(f"æ–‡ä»¶ {file_path} å·²æˆåŠŸåˆ é™¤ã€‚")
        else:
            print(f"æ–‡ä»¶ {file_path} ä¸å­˜åœ¨ï¼Œæ— éœ€åˆ é™¤ã€‚")
    except Exception as e:
        print(f"åˆ é™¤æ–‡ä»¶ {file_path} æ—¶å‘ç”Ÿé”™è¯¯: {e}")



def authorize_MPMate(file_path):
    def run_command_with_file(exe_path, file_path):
        try:
            # æ„å»ºå‘½ä»¤
            command = [exe_path, '-f', file_path]
            
            # è¿è¡Œå‘½ä»¤å¹¶æ•è·è¾“å‡º
            result = subprocess.run(
                command, 
                capture_output=True,  # æ•è·æ ‡å‡†è¾“å‡ºå’Œé”™è¯¯
                text=True,            # å°†è¾“å‡ºè§£ç ä¸ºå­—ç¬¦ä¸²
                check=True            # å¦‚æœå‘½ä»¤å¤±è´¥ä¼šæŠ›å‡ºå¼‚å¸¸
            )
            
            # æ‰“å°å‘½ä»¤çš„è¾“å‡º
            # print("æˆæƒå‘½ä»¤è¾“å‡º:")
            print(result.stdout)
            return result.stdout  # è¿”å›æ ‡å‡†è¾“å‡ºå†…å®¹

        except subprocess.CalledProcessError as e:
            print(f"å‘½ä»¤è¿è¡Œå¤±è´¥: {e}")
            print(f"é”™è¯¯è¾“å‡º: {e.stderr}")
            raise Exception("æˆæƒè½¯ä»¶å‡ºé”™")
            return None

    current_path = os.getcwd()
    # print(f"å½“å‰è·¯å¾„æ˜¯: {current_path}")

    MPMate_exe = ".\\3æ ¡éªŒå·¥å…·\MPMateCli-V1.exe"
    file_path = file_path.replace("/", "\\")
    run_command_with_file(MPMate_exe, file_path)
